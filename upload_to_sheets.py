"""
Google Sheets 업로드 모듈
Excel 파일을 읽어서 Google Spreadsheet로 업로드
"""

import os
import json
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import xlrd  # .xls 파일 지원
import gspread
from google.oauth2.service_account import Credentials

# ===== 동기화 상태 표시 설정 =====
# 동기화 상태를 표시할 셀 위치
# 예: "A1", "Z1", "AA1" 등
# 환경변수 SYNC_STATUS_CELL로 오버라이드 가능
DEFAULT_SYNC_STATUS_CELL = "A1"


class GoogleSheetsUploader:
    def __init__(self, credentials_json: str = None, credentials_dict: dict = None):
        """
        Google Sheets Uploader 초기화

        Args:
            credentials_json: 서비스 계정 JSON 파일 경로
            credentials_dict: 서비스 계정 정보 딕셔너리 (GitHub Secrets용)
        """
        self.scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]

        if credentials_dict:
            credentials = Credentials.from_service_account_info(
                credentials_dict,
                scopes=self.scopes
            )
        elif credentials_json:
            credentials = Credentials.from_service_account_file(
                credentials_json,
                scopes=self.scopes
            )
        else:
            raise ValueError("credentials_json 또는 credentials_dict가 필요합니다.")

        self.client = gspread.authorize(credentials)

    def read_excel(self, file_path: str, sheet_name: str = None) -> list:
        """
        Excel 파일 읽기 (형식 자동 감지)

        Args:
            file_path: Excel 파일 경로
            sheet_name: 읽을 시트 이름 (None이면 첫 번째 시트)

        Returns:
            list: 2D 리스트 (행 데이터)
        """
        print(f"Excel 파일 로딩 중: {file_path}")

        # 파일 시작 부분을 읽어서 형식 감지
        with open(file_path, 'rb') as f:
            header = f.read(100)

        # XML Spreadsheet 형식 감지
        if header.startswith(b'<?xml') or b'<Workbook' in header:
            print("  XML Spreadsheet 형식 감지됨")
            return self._read_xml_spreadsheet(file_path, sheet_name)
        elif file_path.lower().endswith('.xls'):
            return self._read_xls(file_path, sheet_name)
        else:
            return self._read_xlsx(file_path, sheet_name)

    def _read_xml_spreadsheet(self, file_path: str, sheet_name: str = None) -> list:
        """Microsoft Office XML Spreadsheet 형식 읽기"""
        # XML 네임스페이스 정의
        namespaces = {
            'ss': 'urn:schemas-microsoft-com:office:spreadsheet',
            'o': 'urn:schemas-microsoft-com:office:office',
            'x': 'urn:schemas-microsoft-com:office:excel',
        }

        print("  XML 파싱 중...")
        tree = ET.parse(file_path)
        root = tree.getroot()

        # Worksheet 찾기
        worksheets = root.findall('.//ss:Worksheet', namespaces)
        if not worksheets:
            raise ValueError("XML에서 Worksheet를 찾을 수 없습니다.")

        # 시트 선택
        if sheet_name:
            ws = None
            for worksheet in worksheets:
                if worksheet.get('{urn:schemas-microsoft-com:office:spreadsheet}Name') == sheet_name:
                    ws = worksheet
                    break
            if ws is None:
                raise ValueError(f"시트 '{sheet_name}'을 찾을 수 없습니다.")
        else:
            ws = worksheets[0]

        # 데이터 추출
        data = []
        table = ws.find('ss:Table', namespaces)
        if table is None:
            return data

        rows = table.findall('ss:Row', namespaces)
        print(f"  총 {len(rows)} 행 발견")

        for row_idx, row in enumerate(rows):
            row_data = []
            cells = row.findall('ss:Cell', namespaces)

            col_idx = 0
            for cell in cells:
                # ss:Index 속성 처리 (빈 셀 건너뛰기)
                index_attr = cell.get('{urn:schemas-microsoft-com:office:spreadsheet}Index')
                if index_attr:
                    target_idx = int(index_attr) - 1
                    while col_idx < target_idx:
                        row_data.append('')
                        col_idx += 1

                # 셀 데이터 추출
                data_elem = cell.find('ss:Data', namespaces)
                if data_elem is not None and data_elem.text:
                    row_data.append(str(data_elem.text))
                else:
                    row_data.append('')
                col_idx += 1

            data.append(row_data)

            if (row_idx + 1) % 10000 == 0:
                print(f"  {row_idx + 1} 행 처리됨...")

        print(f"Excel 파일 로드 완료: {len(data)} 행")
        return data

    def _read_xls(self, file_path: str, sheet_name: str = None) -> list:
        """xlrd를 사용하여 .xls 파일 읽기"""
        wb = xlrd.open_workbook(file_path)

        if sheet_name:
            ws = wb.sheet_by_name(sheet_name)
        else:
            ws = wb.sheet_by_index(0)

        data = []
        for row_idx in range(ws.nrows):
            row = ws.row_values(row_idx)
            # None 값을 빈 문자열로 변환
            cleaned_row = [str(cell) if cell is not None else '' for cell in row]
            cleaned_row = ['' if cell == 'None' else cell for cell in cleaned_row]
            data.append(cleaned_row)
            if (row_idx + 1) % 10000 == 0:
                print(f"  {row_idx + 1} 행 로드됨...")

        print(f"Excel 파일 로드 완료: {len(data)} 행")
        return data

    def _read_xlsx(self, file_path: str, sheet_name: str = None) -> list:
        """openpyxl을 사용하여 .xlsx 파일 읽기"""
        wb = load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        data = []
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            # None 값을 빈 문자열로 변환
            cleaned_row = [str(cell) if cell is not None else '' for cell in row]
            # 'None' 문자열도 빈 문자열로
            cleaned_row = ['' if cell == 'None' else cell for cell in cleaned_row]
            data.append(cleaned_row)
            row_count += 1
            if row_count % 10000 == 0:
                print(f"  {row_count} 행 로드됨...")

        wb.close()
        print(f"Excel 파일 로드 완료: {len(data)} 행")
        return data

    def update_sync_status(
        self,
        spreadsheet_id: str,
        worksheet_name: str,
        status: str,
        timestamp: str = None,
        cell: str = None
    ) -> bool:
        """
        특정 셀에 동기화 상태 기록

        Args:
            spreadsheet_id: Google Spreadsheet ID
            worksheet_name: 워크시트 이름 (cell에 시트 지정이 없을 때 사용)
            status: 상태 메시지 (예: "동기화 중...", "완료")
            timestamp: 타임스탬프 (None이면 현재 시간)
            cell: 기록할 셀 (None이면 환경변수 또는 기본값 사용)
                 예: "A1", "Z1", "AA1" (현재 시트)
                 또는 "Index!D5" (다른 시트의 특정 셀)

        Returns:
            bool: 성공 여부
        """
        from datetime import datetime, timezone, timedelta

        # 한국 시간대 (UTC+9)
        KST = timezone(timedelta(hours=9))

        # 셀 위치 결정: 매개변수 > 환경변수 > 기본값
        if cell is None:
            cell = os.getenv("SYNC_STATUS_CELL", DEFAULT_SYNC_STATUS_CELL)

        if timestamp is None:
            timestamp = datetime.now(KST).strftime('%Y-%m-%d %H:%M:%S')

        try:
            spreadsheet = self.client.open_by_key(spreadsheet_id)

            # 셀 참조에서 시트 이름과 셀 주소 분리
            # 형식: "SheetName!A1" 또는 "A1"
            if "!" in cell:
                target_sheet_name, cell_address = cell.split("!", 1)
                worksheet = spreadsheet.worksheet(target_sheet_name)
            else:
                # 시트 이름이 없으면 worksheet_name 사용
                worksheet = spreadsheet.worksheet(worksheet_name)
                cell_address = cell

            # 상태 메시지 업데이트
            worksheet.update(range_name=cell_address, values=[[f"{status} ({timestamp})"]])
            print(f"동기화 상태 업데이트: {status} - {cell}")
            return True

        except Exception as e:
            print(f"동기화 상태 업데이트 실패: {e}")
            return False

    def upload_to_spreadsheet(
        self,
        data: list,
        spreadsheet_id: str,
        worksheet_name: str = None,
        clear_existing: bool = True
    ) -> bool:
        """
        데이터를 Google Spreadsheet에 업로드

        Args:
            data: 업로드할 2D 리스트 데이터
            spreadsheet_id: Google Spreadsheet ID
            worksheet_name: 워크시트 이름 (None이면 첫 번째 시트)
            clear_existing: 기존 데이터 삭제 여부

        Returns:
            bool: 성공 여부
        """
        try:
            # Spreadsheet 열기
            spreadsheet = self.client.open_by_key(spreadsheet_id)

            # Worksheet 선택 또는 생성
            if worksheet_name:
                try:
                    worksheet = spreadsheet.worksheet(worksheet_name)
                except gspread.WorksheetNotFound:
                    worksheet = spreadsheet.add_worksheet(
                        title=worksheet_name,
                        rows=len(data) + 1,
                        cols=len(data[0]) if data else 1
                    )
                    print(f"새 워크시트 생성: {worksheet_name}")
            else:
                worksheet = spreadsheet.sheet1

            # 기존 데이터 삭제
            if clear_existing:
                worksheet.clear()
                print("기존 데이터 삭제 완료")

            # 대용량 데이터는 배치로 업로드
            batch_size = 5000  # Google Sheets API 제한 고려
            total_rows = len(data)

            if total_rows <= batch_size:
                worksheet.update(range_name='A1', values=data)
                print(f"업로드 완료: {total_rows} 행")
            else:
                print(f"대용량 데이터 배치 업로드 시작: {total_rows} 행")
                for i in range(0, total_rows, batch_size):
                    batch = data[i:i + batch_size]
                    start_row = i + 1
                    range_name = f'A{start_row}'
                    worksheet.update(range_name=range_name, values=batch)
                    print(f"  배치 업로드: 행 {start_row} ~ {start_row + len(batch) - 1}")

            print(f"업로드 완료: 총 {total_rows} 행")
            return True

        except Exception as e:
            print(f"업로드 중 오류 발생: {str(e)}")
            return False

    def upload_excel_to_sheets(
        self,
        excel_path: str,
        spreadsheet_id: str,
        excel_sheet_name: str = None,
        gsheet_worksheet_name: str = None,
        clear_existing: bool = True
    ) -> bool:
        """
        Excel 파일을 Google Sheets로 업로드 (원스텝)

        Args:
            excel_path: Excel 파일 경로
            spreadsheet_id: Google Spreadsheet ID
            excel_sheet_name: Excel 시트 이름
            gsheet_worksheet_name: Google Sheets 워크시트 이름
            clear_existing: 기존 데이터 삭제 여부

        Returns:
            bool: 성공 여부
        """
        data = self.read_excel(excel_path, excel_sheet_name)
        return self.upload_to_spreadsheet(
            data=data,
            spreadsheet_id=spreadsheet_id,
            worksheet_name=gsheet_worksheet_name,
            clear_existing=clear_existing
        )


def upload_excel_to_google_sheets(
    excel_path: str,
    spreadsheet_id: str,
    credentials_json: str = None,
    credentials_dict: dict = None,
    worksheet_name: str = None
) -> bool:
    """
    Excel 파일을 Google Sheets로 업로드하는 헬퍼 함수

    Args:
        excel_path: Excel 파일 경로
        spreadsheet_id: Google Spreadsheet ID (URL에서 추출)
        credentials_json: 서비스 계정 JSON 파일 경로
        credentials_dict: 서비스 계정 정보 딕셔너리
        worksheet_name: 업로드할 워크시트 이름

    Returns:
        bool: 성공 여부
    """
    uploader = GoogleSheetsUploader(
        credentials_json=credentials_json,
        credentials_dict=credentials_dict
    )

    return uploader.upload_excel_to_sheets(
        excel_path=excel_path,
        spreadsheet_id=spreadsheet_id,
        gsheet_worksheet_name=worksheet_name
    )


if __name__ == "__main__":
    # 테스트용
    from dotenv import load_dotenv
    load_dotenv()

    credentials_json = os.getenv("GOOGLE_CREDENTIALS_PATH")
    spreadsheet_id = os.getenv("GOOGLE_SPREADSHEET_ID")

    # 테스트 Excel 파일이 있다면 실행
    test_excel = os.path.join(os.getcwd(), "downloads", "test.xlsx")

    if os.path.exists(test_excel) and credentials_json and spreadsheet_id:
        success = upload_excel_to_google_sheets(
            excel_path=test_excel,
            spreadsheet_id=spreadsheet_id,
            credentials_json=credentials_json
        )
        print(f"업로드 결과: {'성공' if success else '실패'}")
    else:
        print("테스트를 위한 환경변수 또는 테스트 파일이 없습니다.")
